from openpyxl import load_workbook
import os
def append_to_excel(df, file_path, sheet_name='data_load'):
    if os.path.exists(file_path):
        book = load_workbook(file_path)
        sheet = book[sheet_name] if sheet_name in book.sheetnames else book.create_sheet(sheet_name)

        # Находим первый пустой ряд, проверяя только первую ячейку каждого ряда
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is None:
                startrow = row
                break
        else:
            startrow = sheet.max_row + 1

        # Преобразование DataFrame в список списков
        rows = df.values.tolist()

        # Добавление данных
        for i, row in enumerate(rows, start=startrow):
            for j, value in enumerate(row):
                sheet.cell(row=i, column=j+1, value=value)

        # Сохранение файла
        book.save(file_path)
    else:
        # Если файла нет, создать новый
        df.to_excel(file_path, sheet_name=sheet_name, index=False)